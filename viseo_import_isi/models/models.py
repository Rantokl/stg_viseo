# -*- coding: utf-8 -*-

# from odoo import models, fields, api


# class viseo_import_isi(models.Model):
#     _name = 'viseo_import_isi.viseo_import_isi'
#     _description = 'viseo_import_isi.viseo_import_isi'

#     name = fields.Char()
#     value = fields.Integer()
#     value2 = fields.Float(compute="_value_pc", store=True)
#     description = fields.Text()
#
#     @api.depends('value')
#     def _value_pc(self):
#         for record in self:
#             record.value2 = float(record.value) / 100

from odoo import models, fields
from odoo.exceptions import UserError
import openpyxl
import io
import base64

class ViseoImportISI(models.Model):
    _name = 'viseo_import.isi'
    _description = 'Viseo Import ISI'

    
    nom_prenoms = fields.Char(string='Anaranasyfanampiny*(Nom et Prénoms)')
    num_cin = fields.Char(string='Laharan\'nykarampanomdrom-pirenena*(CIN)')
    nature = fields.Char(string='Anton\'nyfifampivarotana*(Nature)')
    detail_transaction = fields.Char(string='Mombanyfifam-pivarotana(Détail de transaction)')
    date_transaction = fields.Date(string='Datin\'nyfifam-pivarotana*(Date de transaction)')
    montant_transaction = fields.Float(string='Saran\'nyfifampivarotana*(Montant de transaction)', track_visibility='always')
    montant_isi = fields.Float(string='ISImifanarakaaminy*(ISI correspondant)', track_visibility='always')
    province = fields.Char(string='Faritany*(Province)')
    region = fields.Char(string='Faritra*(Région)')
    district = fields.Char(string='District')
    commune = fields.Char(string='Kaominina*(Commune)')
    fokontany = fields.Char(string='Fokontany')

    file_data = fields.Binary(string='Fichier Excel')
    filename = fields.Char(string='Nom du fichier')

    def import_excel_data(self):
        for record in self:
            if not record.file_data:
                raise UserError("Aucun fichier sélectionné.")
            
            try:
                file_data = io.BytesIO(base64.b64decode(record.file_data))
                workbook = openpyxl.load_workbook(file_data)
                worksheet = workbook.active
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))

                if not rows:
                    raise UserError("Le fichier Excel est vide ou mal formaté.")

                records_to_create = []
                for row in rows:
                    if len(row) != 12:
                        raise UserError("Le fichier Excel a un nombre incorrect de colonnes.")

                    records_to_create.append({
                       
                        'nom_prenoms': row[0],
                        'num_cin': row[1],
                        'nature': row[2],
                        'detail_transaction': row[3],
                        'date_transaction': row[4],
                        'montant_transaction': row[5],
                        'montant_isi': row[6],
                        'province': row[7],
                        'region': row[8],
                        'district': row[9],
                        'commune': row[10],
                        'fokontany': row[11],
                        
                    })

                if records_to_create:
                    self.create(records_to_create)

                self.env.cr.commit()
            except Exception as e:
                raise UserError(f"Une erreur est survenue lors de l'importation des données : {e}")


class WizardImportViseoISI(models.TransientModel):
    _name = 'viseo_import_isi.wizard'
    _description = 'Wizard Import ISI'

    file_data = fields.Binary(string='Fichier Excel')
    filename = fields.Char(string='Nom du fichier')

    def import_excel_data_wizard(self):
        for record in self:
            if not record.file_data:
                raise UserError("Aucun fichier sélectionné.")
            
            try:
                file_data = io.BytesIO(base64.b64decode(record.file_data))
                workbook = openpyxl.load_workbook(file_data)
                worksheet = workbook.active
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))

                if not rows:
                    raise UserError("Le fichier Excel est vide ou mal formaté.")

                records_to_create = []
                for row in rows:
                    if len(row) != 12:
                        raise UserError("Le fichier Excel a un nombre incorrect de colonnes.")

                    records_to_create.append({
                        
                        'nom_prenoms': row[0],
                        'num_cin': row[1],
                        'nature': row[2],
                        'detail_transaction': row[3],
                        'date_transaction': row[4],
                        'montant_transaction': row[5],
                        'montant_isi': row[6],
                        'province': row[7],
                        'region': row[8],
                        'district': row[9],
                        'commune': row[10],
                        'fokontany': row[11],
                        
                    })

                if records_to_create:
                    self.env['viseo_import.isi'].create(records_to_create)

                self.env.cr.commit()
            except Exception as e:
                raise UserError(f"Une erreur est survenue lors de l'importation des données : {e}")

