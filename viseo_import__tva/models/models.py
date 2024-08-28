
from odoo import models, fields
from odoo.exceptions import UserError
import openpyxl
import io
import base64

class ViseoImportTVA(models.Model):
    _name = 'viseo_import.tva'
    _description = 'Viseo Import TVA'

    # Définition des champs
    deduc_coll = fields.Char(string='Déductible ou Collectée')
    local_etranger = fields.Char(string='Local ou Etranger')
    num_nif = fields.Char(string='NIF (10 chiffres avec les 0)')
    raison_sociale = fields.Char(string='Raison sociale')
    num_stat = fields.Char(string='Partenaire/Stat')
    adresse = fields.Char(string='Adresse')
    montant_ht = fields.Float(string='Montant HT')
    montant_tva = fields.Float(string='Montant TVA')
    ref_facture = fields.Char(string='Référence facture')
    date_facture = fields.Date(string='Date facture (jj/mm/aaaa)')
    nature = fields.Char(string='Nature')
    libelle_operation = fields.Char(string='Libellé opération')
    date_paiement = fields.Date(string='Date de paiement (jj/mm/aaaa)')
    mois = fields.Integer(string='Mois')
    annee = fields.Integer(string='Année')
    n_dau = fields.Char(string='N DAU')
    code_anex = fields.Char(string='Code Anex')
    code_decl = fields.Char(string='Code décl')
    observation = fields.Char(string='Observation')
    declared = fields.Boolean(string='Déclaré', default=False)

    file_data = fields.Binary(string='Fichier Excel')
    filename = fields.Char(string='Nom du fichier')

    def import_excel_data(self):
        for record in self:
            if not record.file_data:
                raise UserError("Aucun fichier sélectionné.")
            
            try:
                file_data = io.BytesIO(base64.b64decode(record.file_data))
                workbook = openpyxl.load_workbook(file_data)
                worksheet = workbook.active
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))

                if not rows:
                    raise UserError("Le fichier Excel est vide ou mal formaté.")

                records_to_create = []
                for row in rows:
                    if len(row) != 19:
                        raise UserError("Le fichier Excel a un nombre incorrect de colonnes.")

                    records_to_create.append({
                        'deduc_coll': row[0],
                        'local_etranger': row[1],
                        'num_nif': row[2],
                        'raison_sociale': row[3],
                        'num_stat': row[4],
                        'adresse': row[5],
                        'montant_ht': row[6],
                        'montant_tva': row[7],
                        'ref_facture': row[8],
                        'date_facture': row[9],
                        'nature': row[10],
                        'libelle_operation': row[11],
                        'date_paiement': row[12],
                        'mois': row[13],
                        'annee': row[14],
                        'n_dau': row[15],
                        'code_anex': row[16],
                        'code_decl': row[17],
                        'observation': row[18],
                    })

                if records_to_create:
                    self.create(records_to_create)

                self.env.cr.commit()
            except Exception as e:
                raise UserError(f"Une erreur est survenue lors de l'importation des données : {e}")

class EcritureComptableDeclaration(models.Model):
    _inherit = 'account.move.line'

    ref_facture = fields.Char(string='Référence Facture')
    

class WizardImportViseoTVA(models.TransientModel):
    _name = 'viseo_import_tva.wizard'
    _description = 'Wizard Import TVA'

    file_data = fields.Binary(string='Fichier Excel')
    filename = fields.Char(string='Nom du fichier')

    def import_excel_data_wizard(self):
        for record in self:
            if not record.file_data:
                raise UserError("Aucun fichier sélectionné.")
            
            try:
                file_data = io.BytesIO(base64.b64decode(record.file_data))
                workbook = openpyxl.load_workbook(file_data)
                worksheet = workbook.active
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))

                if not rows:
                    raise UserError("Le fichier Excel est vide ou mal formaté.")

                records_to_create = []
                for row in rows:
                    if len(row) != 19:
                        raise UserError("Le fichier Excel a un nombre incorrect de colonnes.")

                    records_to_create.append({
                        'deduc_coll': row[0],
                        'local_etranger': row[1],
                        'num_nif': row[2],
                        'raison_sociale': row[3],
                        'num_stat': row[4],
                        'adresse': row[5],
                        'montant_ht': row[6],
                        'montant_tva': row[7],
                        'ref_facture': row[8],
                        'date_facture': row[9],
                        'nature': row[10],
                        'libelle_operation': row[11],
                        'date_paiement': row[12],
                        'mois': row[13],
                        'annee': row[14],
                        'n_dau': row[15],
                        'code_anex': row[16],
                        'code_decl': row[17],
                        'observation': row[18],
                    })

                if records_to_create:
                    import_records = self.env['viseo_import.tva'].create(records_to_create)

                    # Mise à jour des écritures comptables
                    for record in import_records:
                        if record.ref_facture:
                            # Rechercher les écritures comptables correspondant à la référence
                            # On suppose ici que 'move_id' est le champ liant les lignes comptables aux factures
                            factures = self.env['account.move.line'].search([('ref_facture', '=', record.ref_facture)])
                            if factures:
                                # Mettre à jour les écritures comptables en arrière-plan
                                factures.write({'is_declared': True})

                self.env.cr.commit()
            except Exception as e:
                raise UserError(f"Une erreur est survenue lors de l'importation des données : {e}")


